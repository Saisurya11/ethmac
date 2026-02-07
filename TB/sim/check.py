import os
import openpyxl
import re

book = openpyxl.load_workbook("report.xlsx")
sheet = book["Sheet1"]

f = open("tests.txt", "r")
lists_file = f.readlines()

headers = ["s.no","testname","result","UVM_INFO","UVM_FATAL","UVM_WARNING","UVM_ERROR"]

for col, val in enumerate(headers, start=1):
    sheet.cell(row=1, column=col).value = val

count = 2

for i in lists_file:
    test_name = i.strip()
    sheet.cell(row=count, column=1).value = count - 1
    sheet.cell(row=count, column=2).value = test_name

    test_file = f"{test_name}.txt"

    # ===== FILE EXISTENCE CHECK =====
    if not os.path.isfile(test_file):
        sheet.cell(row=count, column=3).value = "NOT FOUND"
        count += 1
        continue
    # =================================

    with open(test_file, "r", errors="ignore") as file:
        list_file = file.read()

    result = "PASS"

    match = re.search(r".*UVM_FATAL.*:\s*(\d+)", list_file, re.I)
    if match:
        sheet.cell(row=count, column=5).value = match.group(1)
        if int(match.group(1)) > 0:
            result = "FAIL"

    match = re.search(r".*UVM_ERROR.*:\s*(\d+)", list_file, re.I)
    if match:
        sheet.cell(row=count, column=7).value = match.group(1)
        if int(match.group(1)) > 0:
            result = "FAIL"

    match = re.search(r".*UVM_WARNING.*:\s*(\d+)", list_file, re.I)
    if match:
        sheet.cell(row=count, column=6).value = match.group(1)
        if int(match.group(1)) > 0:
            result = "FAIL"

    match = re.search(r".*UVM_INFO.*:\s*(\d+)", list_file, re.I)
    if match:
        sheet.cell(row=count, column=4).value = match.group(1)

    sheet.cell(row=count, column=3).value = result
    count += 1

book.save("report.xlsx")
f.close()

