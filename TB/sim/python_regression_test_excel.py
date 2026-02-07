import os
f=open("tests.txt","r")
list_file=f.readlines()
count=0
for line in list_file:
  #line.strip()
   print(line)
  # if(count<1):
   line_strip=line.strip()
   os.system("vlib work")
   os.system("vmap work work")
   os.system("vlog -ccflags \"-DQUESTA\" C:/uvm_1.2/uvm-1.2-master/src/dpi/uvm_dpi.cc")
   os.system("vlog +incdir+C:/uvm_1.2/uvm-1.2-master/src C:/uvm_1.2/uvm-1.2-master/src/uvm_pkg.sv ../top/top.sv +incdir+C:/uvm_1.2/uvm-1.2-master/src    +incdir+../../rtl    +incdir+../top    +incdir+../wb_proc    +incdir+../wb_mem    +incdir+../mii    +incdir+../phy    +incdir+../sbd +incdir+../reg_model")
   os.system("vopt top +cover=fcbest -o %0s"%(line_strip))
   os.system("vsim -c -voptargs=+acc -coverage %0s +UVM_TESTNAME=%0s -l %0s.txt -do \"coverage save -onexit %0s.ucdb\" -do \"run -all\" -do \"quit\""%(line_strip,line_strip,line_strip,line_strip))
   #os.system("vsim -c -do \"vcov merge final.ucdb %s.ucdb\" -do \"quit\" "%(line_strip))
   count=count+1
f.close()
os.system("vsim -c -do \"vcov merge final.ucdb ethmac_reg_read_test.ucdb ethmac_reg_write_read_test.ucdb ethmac_reg_read_test_reg_model.ucdb ethmac_reg_write_read_reg_model_test.ucdb ethmac_reg_BD_write_read_reg_model_test.ucdb ethmac_10mbps_fd_tx_test.ucdb ethmac_100mbps_fd_tx_test.ucdb ethmac_fd_rx_test.ucdb ethmac_fd_tx_rx_test.ucdb mac_coll_hd_test.ucdb mac_mii_write_ctrl_data.ucdb ctrl_frame_test.ucdb tx_ctrl_frame_test.ucdb\" -do \"quit\"")
os.system("vsim -c -do \"vcover report -details -html final.ucdb\" -do \"quit\"")
print(count)


import openpyxl
import re

book = openpyxl.load_workbook("report.xlsx")
sheet = book["Sheet1"]

f = open("tests.txt", "r")
lists_file = f.readlines()

headers = ["s.no","testname","result","UVM_INFO","UVM_FATAL","UVM_WARNING","UVM_ERROR"]

for col, val in enumerate(headers, start=1):
    sheet.cell(row=1, column=col).value = val

count = 2

for i in lists_file:
    test_name = i.strip()
    sheet.cell(row=count, column=1).value = count - 1
    sheet.cell(row=count, column=2).value = test_name

    test_file = f"{test_name}.txt"

    # ===== FILE EXISTENCE CHECK =====
    if not os.path.isfile(test_file):
        sheet.cell(row=count, column=3).value = "NOT FOUND"
        count += 1
        continue
    # =================================

    with open(test_file, "r", errors="ignore") as file:
        list_file = file.read()

    result = "PASS"

    match = re.search(r".*UVM_SCOREBOARD::PASSED.*",list_file,re.I)
    if match==0:
        print("not matched")
        result = "FAIL"

    match = re.search(r".*UVM_FATAL.*:\s*(\d+)", list_file, re.I)
    if match:
        sheet.cell(row=count, column=5).value = match.group(1)
        if int(match.group(1)) > 0:
            result = "FAIL"

    match = re.search(r".*UVM_ERROR.*:\s*(\d+)", list_file, re.I)
    if match:
        sheet.cell(row=count, column=7).value = match.group(1)
        if int(match.group(1)) > 0:
            result = "FAIL"

    match = re.search(r".*UVM_WARNING.*:\s*(\d+)", list_file, re.I)
    if match:
        sheet.cell(row=count, column=6).value = match.group(1)
        if int(match.group(1)) > 0:
            result = "FAIL"

    match = re.search(r".*UVM_INFO.*:\s*(\d+)", list_file, re.I)
    if match:
        sheet.cell(row=count, column=4).value = match.group(1)

    sheet.cell(row=count, column=3).value = result
    count += 1

book.save("report.xlsx")
f.close()

